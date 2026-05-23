"""
excel_ops.py — append-only operations on the job log.

All writes go through append_rows(), which:
  1. Runs the safety preflight (backup + openable check).
  2. Routes each row to its country sheet.
  3. Deduplicates against existing rows.
  4. Appends — never overwrites.
  5. Applies row formatting (green fill for Selected rows, hyperlink Apply Link).

See job-discovery/references/append-only-safety.md for the rules this enforces.
"""

import os

from safety_checks import preflight, JobLogLocked, BackupFailed


COLUMN_ORDER = [
    "Source", "Selected", "Timestamp", "#", "Job Title", "Company",
    "Location", "Job Type", "Match Score", "Match Justification",
    "Skill Gap", "Salary Range", "Apply Link",
]

GREEN_FILL = "E2EFDA"


def _country_for_location(location, prompt_context_country=None):
    """Resolve a country sheet name from a posting location string.

    See regional-sheet-mapping.md. This is a lookup table for common cases;
    unknown locations fall back to prompt context or a best-effort parse.
    """
    loc = (location or "").lower()
    table = {
        "egypt": "Egypt", "cairo": "Egypt", "alexandria": "Egypt",
        "saudi": "Saudi Arabia", "riyadh": "Saudi Arabia", "jeddah": "Saudi Arabia",
        "denmark": "Denmark", "copenhagen": "Denmark", "aarhus": "Denmark",
        "germany": "Germany", "berlin": "Germany", "munich": "Germany",
        "united kingdom": "United Kingdom", "london": "United Kingdom",
        "uk": "United Kingdom",
        "united states": "United States", "usa": "United States",
        "new york": "United States", "remote, us": "United States",
    }
    for key, sheet in table.items():
        if key in loc:
            return sheet
    if "remote" in loc and prompt_context_country:
        return prompt_context_country
    # Fall back to the last comma-separated token, title-cased.
    if "," in (location or ""):
        return location.split(",")[-1].strip().title()
    return prompt_context_country or "Unsorted"


def _is_duplicate(row, existing_rows):
    """A duplicate matches BOTH company and job title (case-insensitive)."""
    company = (row.get("Company") or "").strip().lower()
    title = (row.get("Job Title") or "").strip().lower()
    for ex in existing_rows:
        if ((ex.get("Company") or "").strip().lower() == company
                and (ex.get("Job Title") or "").strip().lower() == title):
            return True
    return False


def append_rows(job_log_path, rows, interactive=True, prompt_context_country=None):
    """Append job rows to the log, routed to country sheets, deduplicated.

    Args:
        job_log_path: path to Job Listings.xlsx
        rows: list of dicts keyed by COLUMN_ORDER fields
        interactive: True for interactive sessions (hard-stop on lock)
        prompt_context_country: fallback country for "Remote" postings

    Returns a summary dict: {appended, skipped_duplicates, sheets_touched}.

    Raises JobLogLocked / BackupFailed — the caller handles interactive vs.
    unattended behavior (see append-only-safety.md Rules 4 and 5).
    """
    preflight(job_log_path, interactive=interactive)

    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    if os.path.exists(job_log_path):
        wb = load_workbook(job_log_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default sheet

    fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL,
                       fill_type="solid")
    summary = {"appended": 0, "skipped_duplicates": 0, "sheets_touched": set()}

    # Group incoming rows by destination sheet.
    by_sheet = {}
    for row in rows:
        sheet_name = _country_for_location(
            row.get("Location"), prompt_context_country)
        by_sheet.setdefault(sheet_name, []).append(row)

    for sheet_name, sheet_rows in by_sheet.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            existing = _read_existing(ws)
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(COLUMN_ORDER)
            existing = []

        for row in sheet_rows:
            if _is_duplicate(row, existing):
                summary["skipped_duplicates"] += 1
                continue
            values = [row.get(col, "") for col in COLUMN_ORDER]
            ws.append(values)
            r = ws.max_row
            # Apply Link as hyperlink.
            link = row.get("Apply Link", "")
            if link:
                link_col = COLUMN_ORDER.index("Apply Link") + 1
                cell = ws.cell(row=r, column=link_col)
                cell.hyperlink = link
                cell.style = "Hyperlink"
            # Green fill for Selected rows.
            if str(row.get("Selected", "")).strip() == "✓":
                for c in range(1, len(COLUMN_ORDER) + 1):
                    ws.cell(row=r, column=c).fill = fill
            existing.append(row)
            summary["appended"] += 1
            summary["sheets_touched"].add(sheet_name)

    wb.save(job_log_path)
    summary["sheets_touched"] = sorted(summary["sheets_touched"])
    return summary


def _read_existing(ws):
    """Read existing rows of a worksheet into a list of dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    out = []
    for values in rows[1:]:
        out.append(dict(zip(header, values)))
    return out


if __name__ == "__main__":
    print(
        "excel_ops.py is a library module. Import append_rows() from the "
        "job-discovery skill flow. It always runs the safety preflight first."
    )
